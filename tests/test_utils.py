import pytest
from pymasterloadstool import utilities
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet


class TestUtilities:
    def test_max_row_index(self):
        ws = Worksheet(Workbook())
        ws["A3"] = "ABC"
        ws["A4"] = "zxa"
        ws["A5"] = "zxa"
        ws["A6"] = "zxa"
        assert utilities.max_row_index(ws, 3, 500, 1, 1) == 6

    def test_max_column_index(self):
        ws = Worksheet(Workbook())
        ws["A1"] = "ABC"
        ws["B1"] = "zxa"
        ws["C1"] = "zxa"
        ws["D1"] = "zxa"
        assert utilities.max_column_index(ws, 1, 1, 1) == 4

    def test_list_to_str(self):
        test_list = ["4", "5", "acv", "yts", 6, 7.0]
        desired_output = "4, 5, acv, yts, 6, 7.0"
        assert utilities.list_to_str(test_list) == desired_output

    def test_get_key(self):
        test_dictionary = {"a": "100", "b": 101, 3: 45}
        assert utilities.get_key(test_dictionary, "100") == "a"
        assert utilities.get_key(test_dictionary, 101) == "b"
        assert utilities.get_key(test_dictionary, 45) == 3
