from openpyxl import Workbook, load_workbook
from .utilities import max_row_index
from .pyLoad import pyLoad
from .structure import supported_load_types
from .settings import load_sheet_name, cases_sheet_name


class XlsReader:
    """Reads the loads from the excel sheet"""

    def __init__(self, excel_path):
        self.path = excel_path

    def _assign_load_prop(self, load, cell):
        row = cell.row
        load.LCName = self.ws["B" + str(row)].value
        load.LCNumber = self.ws["A" + str(row)].value
        name = self.ws["H" + str(row)].value
        load.Name = name
        load.objects = self.ws["I" + str(row)].value
        load_type = list(supported_load_types.keys())[list(supported_load_types.values()).index(name)]
        if load_type == 0 or load_type == 3:  # nodal force or point load on a bar
            load.FX = self.ws["J" + str(row)].value
            load.FY = self.ws["K" + str(row)].value
            load.FZ = self.ws["L" + str(row)].value
            load.Mx = self.ws["M" + str(row)].value
            load.My = self.ws["N" + str(row)].value
            load.Mz = self.ws["O" + str(row)].value
            load.alfa = self.ws["P" + str(row)].value
            load.beta = self.ws["Q" + str(row)].value
            load.gamma = self.ws["R" + str(row)].value
        if load_type == 26:  # uniform load on a FE element
            load.PX = self.ws["J" + str(row)].value
            load.PY = self.ws["K" + str(row)].value
            load.PZ = self.ws["L" + str(row)].value
            load.cosystem = self.ws["W" + str(row)].value
            load.projected = self.ws["Y" + str(row)].value
        elif load_type == 7:  # self-weight
            load.entirestruc = self.ws["I" + str(row)].value
        elif load_type == 3:  # point load on a bar
            load.cosystem = self.ws["W" + str(row)].value
            load.calcnode = self.ws["V" + str(row)].value
            load.absrel = self.ws["X" + str(row)].value
            load.disX = self.ws["S" + str(row)].value
            load.disY = self.ws["T" + str(row)].value
            load.disZ = self.ws["U" + str(row)].value
        elif load_type == 5:  # uniform load
            load.FX = self.ws["J" + str(row)].value
            load.FY = self.ws["K" + str(row)].value
            load.FZ = self.ws["L" + str(row)].value
            load.Mx = self.ws["M" + str(row)].value
            load.My = self.ws["N" + str(row)].value
            load.Mz = self.ws["O" + str(row)].value
            load.alfa = self.ws["P" + str(row)].value
            load.beta = self.ws["Q" + str(row)].value
            load.gamma = self.ws["R" + str(row)].value
            load.cosystem = self.ws["W" + str(row)].value
            load.absrel = self.ws["X" + str(row)].value
            load.disY = self.ws["T" + str(row)].value
            load.disZ = self.ws["U" + str(row)].value
        elif load_type == 6:  # trapezoidal load (2p)
            load.PX2 = self.ws["J" + str(row)].value
            load.PY2 = self.ws["K" + str(row)].value
            load.PZ2 = self.ws["L" + str(row)].value
            load.PX = self.ws["M" + str(row)].value
            load.PY = self.ws["N" + str(row)].value
            load.PZ = self.ws["O" + str(row)].value
            load.disX = self.ws["T" + str(row)].value
            load.disX2 = self.ws["S" + str(row)].value
            load.alfa = self.ws["P" + str(row)].value
            load.beta = self.ws["Q" + str(row)].value
            load.gamma = self.ws["R" + str(row)].value
            load.projected = self.ws["Y" + str(row)].value
            load.absrel = self.ws["X" + str(row)].value
            # self.ws["V" + str(row)] = load.calcnode
        elif load_type == 69:  # (FE) 2 load on edges
            load.PX = self.ws["J" + str(row)].value
            load.PY = self.ws["K" + str(row)].value
            load.PZ = self.ws["L" + str(row)].value
            load.Mx = self.ws["M" + str(row)].value
            load.My = self.ws["N" + str(row)].value
            load.Mz = self.ws["O" + str(row)].value
            load.gamma = self.ws["R" + str(row)].value
            load.cosystem = self.ws["W" + str(row)].value
        elif load_type == 89:  # "Body forces"
            load.FX = self.ws["J" + str(row)].value
            load.FY = self.ws["K" + str(row)].value
            load.FZ = self.ws["L" + str(row)].value
            load.absrel = self.ws["X" + str(row)].value

    def read_load_data(self):
        loads = []
        start_row = 8
        wb = load_workbook(self.path, data_only=True)
        self.ws = wb[load_sheet_name]
        row_count = max_row_index(self.ws, start_row, max_column=1)
        # clear the range
        load_range = "A" + str(start_row) + ":" + "X" + str(row_count)
        for row in self.ws[load_range]:
            load = pyLoad()
            for cell in row:
                self._assign_load_prop(load, cell)
            loads.append(load)
        return loads

    def read_cases(self):
        # loop through the rows in excel, list unique cases, numbers and natures
        # store them in nested list
        cases = []
        start_row = 7
        wb = load_workbook(self.path, data_only=True)
        self.ws = wb[cases_sheet_name]
        row_count = max_row_index(self.ws, start_row, max_column=1)
        cases_range = "A" + str(start_row) + ":" + "A" + str(row_count)
        for row in self.ws[cases_range]:
            for cell in row:
                # append 0-number,1-name,2-nature int, 3-nonlin, 4-solver, 5-kmatrix, 6-pdelta
                number = self.ws["A" + str(cell.row)].value
                name = self.ws["B" + str(cell.row)].value
                nature = int(self.ws["D" + str(cell.row)].value)
                nonlin = bool(int(self.ws["E" + str(cell.row)].value))
                solver = int(self.ws["F" + str(cell.row)].value)
                kmatrix = bool(int(self.ws["G" + str(cell.row)].value))
                pdelta = bool(int(self.ws["H" + str(cell.row)].value))
                cases.append(
                    [
                        number,
                        name,
                        nature,
                        nonlin,
                        solver,
                        kmatrix,
                        pdelta,
                    ]
                )
        return cases


if __name__ == "__main__":
    path = r"C:\Users\mwo\OneDrive - WoodThilsted Partners\Professional\5_PYTHON\pyMasterLoadsTool\test.xlsx"
    reader = XlsReader(path)
    reader.read_load_data()
