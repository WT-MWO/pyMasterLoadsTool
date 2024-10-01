import os
import clr
from .utilities import max_row_index
from openpyxl.utils import get_column_letter

cwd = os.getcwd()
dll_path = cwd + "\dll\Interop.RobotOM.dll"
clr.AddReference(dll_path)
from RobotOM import *
import RobotOM as rbt


class ContourHandler:
    """Handles the operations on contour points/ contour loads"""

    def __init__(self, ws_points) -> None:
        self.ws_points = ws_points

    def _get_contour_column_number(self, row_id_number: int) -> int:
        """Searches the first row of the 'Contour load data' sheet and returns column number for matching row (load) id"""
        for row in self.ws_points.iter_rows(1, 1, 1, 500):
            for cell in row:
                if cell.data_type == "s":
                    if cell.value == str(row_id_number):
                        return cell.column
                elif cell.data_type == "n":
                    if cell.value == row_id_number:
                        return cell.column

    def _get_contour_points(self, row_id_number: int) -> list[list, list]:
        """ "Reads contour points for matching row (load) id, reutrns them in a list
        [point_number, x, y, z]"""
        points = []
        load_points = []
        point_number = 1
        cell_column = self._get_contour_column_number(row_id_number)
        nsize = max_row_index(self.ws_points, start_row=3, min_column=cell_column, max_column=cell_column)
        contours_range = get_column_letter(cell_column) + "3" + ":" + get_column_letter(cell_column) + str(nsize)
        for row in self.ws_points[contours_range]:
            for cell in row:
                # x = cell.value
                # y = cell.offset(0, 1)
                # z = cell.offset(0, 2)
                points.append(
                    [point_number, float(cell.value), float(cell.offset(0, 1).value), float(cell.offset(0, 2).value)]
                )
                load_points.append([point_number, cell.offset(0, 3).value])
                point_number += 1
        return points, load_points[:3]

    def _assign_contour_points(
        self,
        load_record: rbt.IRobotLoadRecordInContour,
        contour_index: int,
        is_3p: bool = False,
    ) -> None:
        """Assigns contour points and the 3p contour loads corners, A, B, C"""
        points = self._get_contour_points(contour_index)[0]
        load_points = self._get_contour_points(contour_index)[1]
        nsize = len(points)
        load_record.SetValue(13, nsize)  # assign size for point array I_ICRV_NPOINTS
        for point in points:
            load_record.SetContourPoint(point[0], point[1], point[2], point[3])
        if is_3p:
            corner_number = 1
            for lp in load_points:
                # the string from excel is like: A(-1, 2.5, 0.5, 0.0)
                load_text = lp[1][1:].replace("(", "").replace(")", "").split(",")  # remove characters
                load_lst = [float(i) for i in load_text]  # convert to numbers
                load_record.SetPoint(corner_number, load_lst[1], load_lst[2], load_lst[3])
                corner_number += 1

    def _read_contour_points(self, rec: rbt.IRobotLoadRecord) -> list[list, tuple]:
        """Reads contour points.
        Parameters: rec: IRobotLoadRecord
        Returns: list[contour_points, vector]"""
        npoints = int(rec.GetValue(13))
        rec_contour = rbt.IRobotLoadRecordInContour(rec)
        contour_points = []
        for p in range(1, npoints + 1):
            contour_points.append(rec_contour.GetContourPoint(p))
        vector = rec_contour.GetVector()
        return [contour_points, vector]
