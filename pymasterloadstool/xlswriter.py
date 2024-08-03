from openpyxl import Workbook, load_workbook
from .pyLoad import invalid_load
from pymasterloadstool import utilities
from .settings import load_sheet_name, range_to_clear


class XlsWriter:
    """Writes the loads to the excel sheet"""

    def __init__(self, excel_path):
        self.path = excel_path

    def write_data(self, data):
        start_row = 8
        wb = load_workbook(self.path)
        self.ws = wb[load_sheet_name]
        # clear the range
        for r in range_to_clear:
            utilities.clear_range(self.ws, r)
        # write the values
        for load in data:
            if not invalid_load(load):
                self._write_load(load, start_row)
                start_row += 1
        wb.save("test.xlsx")

    def _list_to_str(self, list):
        return ", ".join(repr(e).replace("'", "") for e in list)

    def _write_load(self, load, row):
        """writes a load into excel"""
        # string = "B" + str(row)
        self.ws["B" + str(row)] = load.LCName
        self.ws["A" + str(row)] = load.LCNumber
        self.ws["H" + str(row)] = load.Name
        self.ws["I" + str(row)] = self._list_to_str(load.objects)
        if load.type == 0 or load.type == 3:  # nodal force or point load on a bar
            self.ws["J" + str(row)] = load.FX
            self.ws["K" + str(row)] = load.FY
            self.ws["L" + str(row)] = load.FZ
            self.ws["M" + str(row)] = load.Mx
            self.ws["N" + str(row)] = load.My
            self.ws["O" + str(row)] = load.Mz
            self.ws["P" + str(row)] = load.alfa
            self.ws["Q" + str(row)] = load.beta
            self.ws["R" + str(row)] = load.gamma
            # self.ws["V" + str(row)] = load.calcnode
        if load.type == 26:  # uniform load on a FE element
            self.ws["J" + str(row)] = load.PX
            self.ws["K" + str(row)] = load.PY
            self.ws["L" + str(row)] = load.PZ
            self.ws["W" + str(row)] = load.cosystem
            self.ws["Y" + str(row)] = load.projected
        elif load.type == 7:  # self-weight
            self.ws["I" + str(row)] = load.entirestruc
        elif load.type == 3:  # point load on a bar
            self.ws["W" + str(row)] = load.cosystem
            self.ws["V" + str(row)] = load.calcnode
            self.ws["X" + str(row)] = load.absrel
            self.ws["S" + str(row)] = load.disX
            self.ws["T" + str(row)] = load.disY
            self.ws["U" + str(row)] = load.disZ
        elif load.type == 5:  # uniform load
            self.ws["J" + str(row)] = load.FX
            self.ws["K" + str(row)] = load.FY
            self.ws["L" + str(row)] = load.FZ
            self.ws["M" + str(row)] = load.Mx
            self.ws["N" + str(row)] = load.My
            self.ws["O" + str(row)] = load.Mz
            self.ws["P" + str(row)] = load.alfa
            self.ws["Q" + str(row)] = load.beta
            self.ws["R" + str(row)] = load.gamma
            self.ws["W" + str(row)] = load.cosystem
            self.ws["X" + str(row)] = load.absrel
            self.ws["T" + str(row)] = load.disY
            self.ws["U" + str(row)] = load.disZ
        elif load.type == 6:  # trapezoidal load (2p)
            self.ws["J" + str(row)] = load.PX2
            self.ws["K" + str(row)] = load.PY2
            self.ws["L" + str(row)] = load.PZ2
            self.ws["M" + str(row)] = load.PX
            self.ws["N" + str(row)] = load.PY
            self.ws["O" + str(row)] = load.PZ
            self.ws["T" + str(row)] = load.disX
            self.ws["S" + str(row)] = load.disX2
            self.ws["P" + str(row)] = load.alfa
            self.ws["Q" + str(row)] = load.beta
            self.ws["R" + str(row)] = load.gamma
            self.ws["Y" + str(row)] = load.projected
            self.ws["X" + str(row)] = load.absrel
            # self.ws["V" + str(row)] = load.calcnode
        elif load.type == 69:  # (FE) 2 load on edges
            self.ws["J" + str(row)] = load.PX
            self.ws["K" + str(row)] = load.PY
            self.ws["L" + str(row)] = load.PZ
            self.ws["M" + str(row)] = load.Mx
            self.ws["N" + str(row)] = load.My
            self.ws["O" + str(row)] = load.Mz
            self.ws["R" + str(row)] = load.gamma
            self.ws["W" + str(row)] = load.cosystem
        elif load.type == 89:  # "Body forces"
            self.ws["J" + str(row)] = load.FX
            self.ws["K" + str(row)] = load.FY
            self.ws["L" + str(row)] = load.FZ
            self.ws["X" + str(row)] = load.absrel


if __name__ == "__main__":
    path = r"C:\Users\mwo\OneDrive - WoodThilsted Partners\Professional\5_PYTHON\pyMasterLoadsTool\pyMaster_loads_tool.xlsx"
    writer = XlsWriter(path)
