import math
import clr
from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.utils import get_column_letter
import os
import time

# import ctypes

cwd = os.getcwd()
dll_path = cwd + "\dll\Interop.RobotOM.dll"
clr.AddReference(dll_path)
from RobotOM import *
import RobotOM as rbt

from .structure import Structure, supported_load_types, combination_type, supported_cases_nature
from .enums import cases_nature, case_analize_type
from .utilities import max_row_index, get_key, max_column_index
from .settings import load_sheet_name, cases_sheet_name, combinations_sheet_name, points_sheet_name

M = 1000  # multiplier to get N or Pa
deg_to_rad = math.pi / 180


class Exporter(Structure):
    """Exports the loads from the excel to the Robot model"""

    def __init__(self, app: rbt.RobotApplicationClass, path: str) -> None:
        super().__init__(app)
        self.path = path
        self.wb = load_workbook(self.path, data_only=True)
        self.ws_points = self.wb[points_sheet_name]

    def _del_all_cases(self):
        """Deletes all simple loadcases in the model"""
        case_selection = self.structure.Selections.CreatePredefined(
            rbt.IRobotPredefinedSelection.I_PS_CASE_SIMPLE_CASES
        )
        self.cases.DeleteMany(case_selection)

    def _del_all_combinations(self):
        """Deletes all combinations in the model"""
        comb_selection = self.structure.Selections.CreatePredefined(
            rbt.IRobotPredefinedSelection.I_PS_CASE_COMBINATIONS
        )
        self.cases.DeleteMany(comb_selection)

    def _read_cases(self):
        """Reads loadcases from Excel and stores the input in a list"""
        # loop through the rows in excel, list unique cases, numbers and natures
        # store them in nested list
        cases = []
        start_row = 7
        # wb = load_workbook(self.path, data_only=True)
        self.ws = self.wb[cases_sheet_name]
        row_count = max_row_index(self.ws, start_row, max_column=1)
        cases_range = "A" + str(start_row) + ":" + "A" + str(row_count)
        for row in self.ws[cases_range]:
            for cell in row:
                # append 0-number,1-name,2-nature int, 3-nonlin, 4-solver, 5-kmatrix, 6-pdelta
                number = self.ws["A" + str(cell.row)].value
                name = self.ws["B" + str(cell.row)].value
                # nature = int(self.ws["D" + str(cell.row)].value)
                nature = get_key(supported_cases_nature, self.ws["C" + str(cell.row)].value)
                nonlin = bool(int(self.ws["D" + str(cell.row)].value))
                solver = int(self.ws["E" + str(cell.row)].value)
                kmatrix = bool(int(self.ws["F" + str(cell.row)].value))
                pdelta = bool(int(self.ws["G" + str(cell.row)].value))
                auxilary = bool(int(self.ws["H" + str(cell.row)].value))
                cases.append([number, name, nature, nonlin, solver, kmatrix, pdelta, auxilary])
        return cases

    def _apply_load_cases(self, cases: list[int, str, int, int, int, int, int, int]) -> None:
        """Create load cases in the model
        Paremeters:
        cases: list[0-number,1-name,2-nature int, 3-nonlin, 4-solver, 5-kmatrix, 6-pdelta, 7-auxilary]"""
        for c in cases:
            number = c[0]
            name = c[1]
            nature = cases_nature[c[2]]
            solver = case_analize_type[c[4]]
            case = self.structure.Cases.CreateSimple(number, name, nature, solver)
            if c[7] is True:
                case.IsAuxiliary = True
            # set parameters if solver is non-linear or buckling
            if c[4] == 2:
                params = rbt.IRobotNonlinearAnalysisParams(case.GetAnalysisParams())
                if c[5] is True:
                    params.MatrixUpdateAfterEachIteration = True
                else:
                    params.MatrixUpdateAfterEachIteration = False
                if c[6] is True:
                    params.PDelta = True
                else:
                    params.PDelta = False
                case.SetAnalysisParams(params)

            if c[4] == 4:
                params = rbt.IRobotBucklingAnalysisParams(case.GetAnalysisParams())
                if c[5] is True:
                    params.MatrixUpdateAfterEachIteration = True
                else:
                    params.MatrixUpdateAfterEachIteration = False
                if c[6] is True:
                    params.PDelta = True
                else:
                    params.PDelta = False
                case.SetAnalysisParams(params)

    def _assign_cosystem(self, cosystem_str: str) -> int:
        if cosystem_str == "global":
            return 0
        else:
            return 1

    def _assign_relabs(self, abrel_string: str) -> int:
        if abrel_string == "absolute":
            return 0
        else:
            return 1

    def _assign_comb_nature(self, type: str) -> int:
        if type == 2:
            return 5
        else:
            return 1

    def _get_contour_column_number(self, row_id_number: int) -> int:
        """Searches the first row of the Contour load data sheet and returns column number for matching row (load) id"""
        for row in self.ws_points.iter_rows(1, 1, 1, 500):
            for cell in row:
                if cell.data_type == "s":
                    if cell.value == str(row_id_number):
                        # print(cell.data_type)
                        return cell.column
                elif cell.data_type == "n":
                    if cell.value == row_id_number:
                        # print(cell.data_type)
                        return cell.column

    def _get_contour_points(self, row_id_number: int) -> list[list, list]:
        """ "Reads contour points for matching row (load) id, reutrns them in a list
        [point_number, x, y, z]"""
        points = []
        load_points = []
        point_number = 1
        cell_column = self._get_contour_column_number(row_id_number)
        nsize = max_row_index(self.ws_points, start_row=3, min_column=cell_column, max_column=cell_column)
        contours_range = get_column_letter(cell_column) + "3" + ":" + get_column_letter(cell_column) + str(nsize)
        for row in self.ws_points[contours_range]:
            for cell in row:
                # x = cell.value
                # y = cell.offset(0, 1)
                # z = cell.offset(0, 2)
                points.append(
                    [point_number, float(cell.value), float(cell.offset(0, 1).value), float(cell.offset(0, 2).value)]
                )
                load_points.append([point_number, cell.offset(0, 3).value])
                point_number += 1
        return points, load_points[:3]

    def _assign_contour_points(
        self, load_record: rbt.IRobotLoadRecordInContour, row_id_number: int, is_3p: bool = False
    ) -> None:
        """Assigns contour points and the 3p contour loads corners, A, B, C"""
        points = self._get_contour_points(row_id_number)[0]
        load_points = self._get_contour_points(row_id_number)[1]
        nsize = len(points)
        load_record.SetValue(13, nsize)  # assign size for point array I_ICRV_NPOINTS
        for point in points:
            load_record.SetContourPoint(point[0], point[1], point[2], point[3])
        if is_3p:
            corner_number = 1
            for lp in load_points:
                # the string from excel is like: A(-1, 2.5, 0.5, 0.0)
                load_text = lp[1][1:].replace("(", "").replace(")", "").split(",")  # remove characters
                load_lst = [float(i) for i in load_text]  # convert to numbers
                load_record.SetPoint(corner_number, load_lst[1], load_lst[2], load_lst[3])
                corner_number += 1

    def _assign_loads(self, cell: Cell) -> None:
        """Auxilary function to export one load to the Robot model"""
        row = cell.row
        name = self.ws["H" + str(row)].value
        if name == "load 3p on contour":
            load_type = 28
        else:
            load_type = list(supported_load_types.keys())[list(supported_load_types.values()).index(name)]
        case_number = self.ws["A" + str(row)].value
        objects = self.ws["I" + str(row)].value
        case = rbt.IRobotSimpleCase(self.structure.Cases.Get(case_number))
        if load_type == 7:  # self-weight
            record_index = case.Records.New(rbt.IRobotLoadRecordType(7))
            record = rbt.IRobotLoadRecord(case.Records.Get(record_index))
            load_factor_X = self.ws["J" + str(row)].value
            load_factor_Y = self.ws["K" + str(row)].value
            load_factor_Z = self.ws["L" + str(row)].value
            if load_factor_X != 0:
                sign = math.copysign(1, load_factor_X)
                record.SetValue(0, sign)
                record.SetValue(3, abs(load_factor_X))
            elif load_factor_Y != 0:
                sign = math.copysign(1, load_factor_Y)
                record.SetValue(1, sign)
                record.SetValue(3, abs(load_factor_Y))
            elif load_factor_Z != 0:
                sign = math.copysign(1, load_factor_Z)
                record.SetValue(2, sign)
                record.SetValue(3, abs(load_factor_Z))
            record.Objects.FromText(str(objects))
        elif load_type == 0:  # nodal force
            record_index = case.Records.New(rbt.IRobotLoadRecordType(0))
            record = case.Records.Get(record_index)
            record.Objects.FromText(str(objects))
            record.SetValue(0, self.ws["J" + str(row)].value * M)  # Fx
            record.SetValue(1, self.ws["K" + str(row)].value * M)  # Fy
            record.SetValue(2, self.ws["L" + str(row)].value * M)  # Fz
            record.SetValue(3, self.ws["M" + str(row)].value * M)  # Mx
            record.SetValue(4, self.ws["N" + str(row)].value * M)  # My
            record.SetValue(5, self.ws["O" + str(row)].value * M)  # Mz
            record.SetValue(8, self.ws["P" + str(row)].value * deg_to_rad)  # alfa
            record.SetValue(9, self.ws["Q" + str(row)].value * deg_to_rad)  # beta
            record.SetValue(10, self.ws["R" + str(row)].value * deg_to_rad)  # gamma
        elif load_type == 5:  # uniform load
            record_index = case.Records.New(rbt.IRobotLoadRecordType(5))
            record = case.Records.Get(record_index)
            record.Objects.FromText(str(objects))
            record.SetValue(0, self.ws["J" + str(row)].value * M)  # Fx
            record.SetValue(1, self.ws["K" + str(row)].value * M)  # Fy
            record.SetValue(2, self.ws["L" + str(row)].value * M)  # Fz
            record.SetValue(3, self.ws["M" + str(row)].value * M)  # Mx
            record.SetValue(4, self.ws["N" + str(row)].value * M)  # My
            record.SetValue(5, self.ws["O" + str(row)].value * M)  # Mz
            record.SetValue(8, self.ws["P" + str(row)].value * deg_to_rad)  # alfa
            record.SetValue(9, self.ws["Q" + str(row)].value * deg_to_rad)  # beta
            record.SetValue(10, self.ws["R" + str(row)].value * deg_to_rad)  # gamma
            record.SetValue(11, self._assign_cosystem(self.ws["W" + str(row)].value))  # cosystem
            record.SetValue(13, self._assign_relabs(self.ws["Y" + str(row)].value))  # relabs
            record.SetValue(21, self.ws["T" + str(row)].value)  # dis_y
            record.SetValue(22, self.ws["U" + str(row)].value)  # diz_z
        elif load_type == 3:  # point load on a bar
            record_index = case.Records.New(rbt.IRobotLoadRecordType(3))
            record = case.Records.Get(record_index)
            record.Objects.FromText(str(objects))
            record.SetValue(0, self.ws["J" + str(row)].value * M)  # Fx
            record.SetValue(1, self.ws["K" + str(row)].value * M)  # Fy
            record.SetValue(2, self.ws["L" + str(row)].value * M)  # Fz
            record.SetValue(3, self.ws["M" + str(row)].value * M)  # Mx
            record.SetValue(4, self.ws["N" + str(row)].value * M)  # My
            record.SetValue(5, self.ws["O" + str(row)].value * M)  # Mz
            record.SetValue(8, self.ws["P" + str(row)].value * deg_to_rad)  # alfa
            record.SetValue(9, self.ws["Q" + str(row)].value * deg_to_rad)  # beta
            record.SetValue(10, self.ws["R" + str(row)].value * deg_to_rad)  # gamma
            record.SetValue(11, self._assign_cosystem(self.ws["W" + str(row)].value))  # cosystem
            record.SetValue(12, self.ws["V" + str(row)].value)  # calcnode
            record.SetValue(13, self._assign_relabs(self.ws["X" + str(row)].value))  # relabs
            record.SetValue(6, self.ws["S" + str(row)].value)  # disX
            record.SetValue(21, self.ws["T" + str(row)].value)  # disY
            record.SetValue(22, self.ws["U" + str(row)].value)  # disZ
        elif load_type == 26:  # uniform load on a FE element
            record_index = case.Records.New(rbt.IRobotLoadRecordType(26))
            record = case.Records.Get(record_index)
            record.Objects.FromText(str(objects))
            record.SetValue(0, self.ws["J" + str(row)].value * M)  # Px
            record.SetValue(1, self.ws["K" + str(row)].value * M)  # Py
            record.SetValue(2, self.ws["L" + str(row)].value * M)  # Pz
            record.SetValue(11, self._assign_cosystem(self.ws["W" + str(row)].value))  # cosystem
            record.SetValue(12, self.ws["Y" + str(row)].value)  # projected
        elif load_type == 6:  # trapezoidal load (2p)
            record_index = case.Records.New(rbt.IRobotLoadRecordType(6))
            record = case.Records.Get(record_index)
            record.Objects.FromText(str(objects))
            record.SetValue(0, self.ws["J" + str(row)].value)  # Px2
            record.SetValue(1, self.ws["K" + str(row)].value)  # Py2
            record.SetValue(2, self.ws["L" + str(row)].value)  # Pz2
            record.SetValue(3, self.ws["M" + str(row)].value)  # Px
            record.SetValue(4, self.ws["N" + str(row)].value)  # Py
            record.SetValue(5, self.ws["O" + str(row)].value)  # Pz
            record.SetValue(7, self.ws["T" + str(row)].value)  # dix
            record.SetValue(6, self.ws["S" + str(row)].value)  # disX2
            record.SetValue(8, self.ws["P" + str(row)].value * deg_to_rad)  # alfa
            record.SetValue(9, self.ws["Q" + str(row)].value * deg_to_rad)  # beta
            record.SetValue(10, self.ws["R" + str(row)].value * deg_to_rad)  # gamma
            record.SetValue(12, self.ws["Y" + str(row)].value)  # projected
            record.SetValue(13, self._assign_relabs(self.ws["X" + str(row)].value))
        elif load_type == 69:  # (FE) 2 load on edges
            record_index = case.Records.New(rbt.IRobotLoadRecordType(69))
            record = case.Records.Get(record_index)
            record.Objects.FromText(str(objects))
            record.SetValue(0, self.ws["J" + str(row)].value * M)  # Px
            record.SetValue(1, self.ws["K" + str(row)].value * M)  # Py
            record.SetValue(2, self.ws["L" + str(row)].value * M)  # Pz
            record.SetValue(3, self.ws["M" + str(row)].value * M)  # Mx
            record.SetValue(4, self.ws["N" + str(row)].value * M)  # My
            record.SetValue(5, self.ws["O" + str(row)].value * M)  # Mz
            record.SetValue(6, self.ws["R" + str(row)].value * deg_to_rad)  # gamma
            record.SetValue(11, self._assign_cosystem(self.ws["W" + str(row)].value))  # localsystem
        elif (
            load_type == 89
        ):  # "Body forces" # this is not used anymore TODO: add a messagebox warning when this is detected
            # record_index = case.Records.New(rbt.IRobotLoadRecordType(89, True))
            # record = case.Records.Get(record_index)
            # # record.Objects.FromText(str(objects))
            # record.SetValue(0, self.ws["J" + str(row)].value * M)  # Px
            # record.SetValue(1, self.ws["K" + str(row)].value * M)  # Py
            # record.SetValue(2, self.ws["L" + str(row)].value * M)  # Pz
            # record.SetValue(13, self._assign_relabs(self.ws["X" + str(row)].value))  # relabs
            pass
        elif load_type == 28:  # load on contour
            record_index = case.Records.New(rbt.IRobotLoadRecordType(28))
            record = rbt.IRobotLoadRecordInContour(case.Records.Get(record_index))
            record.Objects.FromText(str(objects))
            record.SetVector(0, 0, -1)  # vector
            # record.SetValue(
            #     rbt.IRobotInContourRecordValues.I_ICRV_AUTO_DETECT_OBJECTS, 1
            # )  # set I_ICRV_AUTO_DETECT_OBJECTS to True, this does not work properly, TODO: investigate

            if name != "load 3p on contour":
                record.SetValue(0, self.ws["M" + str(row)].value * M)  # Px
                record.SetValue(1, self.ws["N" + str(row)].value * M)  # Py
                record.SetValue(2, self.ws["O" + str(row)].value * M)  # Pz
                is_3p = False
            else:
                record.SetValue(3, self.ws["J" + str(row)].value * M)  # Px2
                record.SetValue(4, self.ws["K" + str(row)].value * M)  # Py2
                record.SetValue(5, self.ws["L" + str(row)].value * M)  # Pz2
                record.SetValue(0, self.ws["M" + str(row)].value * M)  # Px1
                record.SetValue(1, self.ws["N" + str(row)].value * M)  # Py1
                record.SetValue(2, self.ws["O" + str(row)].value * M)  # Pz1
                record.SetValue(6, self.ws["S" + str(row)].value * M)  # Px3
                record.SetValue(7, self.ws["T" + str(row)].value * M)  # Py3
                record.SetValue(8, self.ws["U" + str(row)].value * M)  # Pz3
                is_3p = True
            self._assign_contour_points(record, row, is_3p)
        elif load_type == 22:  # load planar trapez TODO: to be implemented
            # record_index = case.Records.New(rbt.IRobotLoadRecordType(22))
            # record = case.Records.Get(record_index)
            # record.Objects.FromText(objects)
            pass
        elif load_type == 8:  # thermal TODO: to be implemented
            # record_index = case.Records.New(rbt.IRobotLoadRecordType(8))
            # record = case.Records.Get(record_index)
            # record.Objects.FromText(objects)
            pass

    def _apply_loads(self) -> None:
        """Applies the loads to the Robot model"""
        start_row = 8
        self.ws = self.wb[load_sheet_name]
        row_count = max_row_index(self.ws, start_row, max_column=1)
        # clear the range
        load_range = "A" + str(start_row) + ":" + "A" + str(row_count)
        for row in self.ws[load_range]:
            for cell in row:
                self._assign_loads(cell)

    def _read_combinations(self) -> None:
        """Read excel input and store it in a list"""
        self.ws_comb = self.wb[combinations_sheet_name]
        start_row = 7
        row_count = max_row_index(self.ws_comb, start_row, max_column=1)
        max_col = max_column_index(self.ws_comb, start_col=1, min_row=3, max_row=3)
        combinations_range = "A" + str(start_row) + ":" + "A" + str(row_count)
        combinations = []
        for row in self.ws_comb[combinations_range]:
            for cell in row:
                # append 0-number,1-name,2-nature int, 3-nonlin, 4-solver, 5-kmatrix, 6-pdelta
                comb_number = self.ws_comb["A" + str(cell.row)].value
                comb_name = self.ws_comb["B" + str(cell.row)].value
                comb_type = get_key(combination_type, self.ws_comb["C" + str(cell.row)].value)
                comb_nature = self._assign_comb_nature(comb_type)
                if self.ws_comb["D" + str(cell.row)].value == 0:
                    comb_analize_type = 0
                    kmatrix = 0
                    pdelta = 0
                else:
                    comb_analize_type = self.ws_comb["E" + str(cell.row)].value
                    kmatrix = self.ws_comb["F" + str(cell.row)].value
                    pdelta = self.ws_comb["G" + str(cell.row)].value
                factors = self._read_factors(row_index=cell.row, min_col=8, max_col=max_col)
            combinations.append(
                [comb_number, comb_name, comb_type, comb_nature, comb_analize_type, kmatrix, pdelta, factors]
            )
        return combinations

    def _read_factors(self, row_index: int, min_col: int, max_col: int) -> list:
        "Reads Excel factor input table and stores the pairs in the list"
        factors = []
        for row in self.ws_comb.iter_rows(min_row=row_index, max_row=row_index, min_col=min_col, max_col=max_col):
            for cell in row:
                if cell.value is not None and cell.value != 0:
                    factor = cell.value
                    loadcase_number = self.ws_comb[cell.column_letter + "4"].value
                    factors.append({loadcase_number: factor})
        return factors

    def _apply_combinations(self, combinations: list[int, str, int, int, int, int, int, int]) -> None:
        """Auxilary function to export one combination to the Robot model"""
        for c in combinations:
            comb_number = c[0]
            comb_name = c[1]
            comb_type = rbt.IRobotCombinationType(c[2])
            comb_nature = rbt.IRobotCaseNature(c[3])
            comb_analize_type = rbt.IRobotCaseAnalizeType(c[4])
            kmatrix = c[5]
            pdelta = c[6]
            factors = c[7]
            cases = self.structure.Cases
            combination = cases.CreateCombination(comb_number, comb_name, comb_type, comb_nature, comb_analize_type)
            if comb_analize_type == 1:
                params = rbt.IRobotNonlinearAnalysisParams(combination.GetAnalysisParams())
                if kmatrix == 1:
                    params.MatrixUpdateAfterEachIteration = True
                if pdelta == 1:
                    params.PDelta = True
                combination.SetAnalysisParams(params)
            # Apply factors
            case_factor_mng = combination.CaseFactors
            for f in factors:
                for key in f.keys():
                    case_factor_mng.New(key, f[key])

    def _export_load_and_cases(self) -> None:
        """Auxilary function for export of loads and loadcases"""
        cases = self._read_cases()
        self._del_all_cases()
        self._apply_load_cases(cases)
        self._apply_loads()

    def _export_combinations(self) -> None:
        """Auxilary function for export load combinations"""
        combinations = self._read_combinations()
        self._del_all_combinations()
        self._apply_combinations(combinations=combinations)

    def export_load_cases_combinations(self, export_loads, export_combinations) -> None:
        """Main function to export"""
        start_time = time.time()
        print("Starting....")
        if export_loads == 1:
            self._export_load_and_cases()
            print("Loads and cases exported. " + str(time.time() - start_time))
        if export_combinations == 1:
            print
            self._export_combinations()
            print("Combinations exported. " + str(time.time() - start_time))
            print("Finished.")
