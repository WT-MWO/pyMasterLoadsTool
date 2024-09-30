from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from typing import Any


def clear_range(ws, range: str) -> None:
    """Clears a cell content in given range
    Slow, do not use. OBSOLETE"""
    cell_range = ws[range]
    for row in cell_range:
        for cell in row:
            cell.value = None


def get_key(dict: dict, val: Any) -> Any:
    """Returns a key from dictionary from given value"""
    for key, value in dict.items():
        if val == value:
            return key


def clear_range2(ws: Worksheet, min_row: int = 1, max_row: int = 1, min_col: int = 1, max_col: int = 1) -> None:
    """Clears a cell content in given range Fast runtime"""
    for row in ws.iter_rows(min_row, max_row, min_col, max_col):
        for cell in row:
            cell.value = None


def max_row_index(ws: Worksheet, start_row: int = 1, max_row: int = 1, min_column: int = 1, max_column: int = 1) -> int:
    """Returns last row index containing data."""
    if start_row == 1:
        count = 1
    else:
        count = start_row - 1
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row, min_col=min_column, max_col=max_column):
        for cell in row:
            if cell.value is not None:
                count += 1
    return count


def max_column_index(ws: Worksheet, min_column: int = 1, min_row: int = 1, max_row: int = 1) -> int:
    """Returns last column index containing data."""
    if min_column == 1:
        count = 1
    else:
        count = min_column - 1
    for col in ws.iter_cols(min_row=min_row, min_col=min_column, max_col=ws.max_column, max_row=max_row):
        for cell in col:
            if cell.value is not None and len(str(cell.value)) > 0:
                count += 1
    return count


def write_list_to_file(list: list, path: str, filename: str) -> None:
    """Writes each item in list to new line in .txt file."""
    # Fix this function to write plain text without
    # parantesis always the same if its a set or list
    with open(path + filename, "w") as file:
        for item in list:
            file.write("{}\n".format(str(item)))


def list_to_str(list: list) -> str:
    """Returns string formatted as: 'a, b, 1, c,' from given list."""
    return ", ".join(repr(e).replace("'", "") for e in list)
