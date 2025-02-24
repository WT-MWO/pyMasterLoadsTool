import tkinter.messagebox
import clr
import math
import time
import os
import tkinter

# from .pyLoad import pyLoad, missing_msg
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

cwd = os.getcwd()
dll_path = cwd + "\dll\Interop.RobotOM.dll"
clr.AddReference(dll_path)
from RobotOM import *
import RobotOM as rbt

from .structure import Structure, supported_load_types, supported_cases_nature, combination_type
from .settings import load_sheet_name, cases_sheet_name, combinations_sheet_name, points_sheet_name
from pymasterloadstool import utilities
from .enums import supported_analize_type

U = 1000  # divider to get kN
R = 2  # rounding
rad_to_deg = 180 / math.pi
missing_msg = "MISSING!!"

# TODO: Import loadcase settings nlin etc.


class Importer(Structure):
    """Reads the loads and combinations from the model, imports them as a list of pyLoad objects."""

    def __init__(self, app: rbt.RobotApplicationClass, path: str) -> None:
        super().__init__(app)
        # self.supported_load_types = supported_load_types
        self.path = path
        self.wb = load_workbook(self.path)
        self.ws_points = self.wb[points_sheet_name]

    def messagebox(self, title: str, text: str) -> None:
        """Warning type message box, closed with OK."""
        root = tkinter.Tk()
        root.withdraw()
        tkinter.messagebox.showwarning(title, text)
        root.destroy()

    @staticmethod
    def _read_cosystem(cosys: int) -> str:
        """Reads if load is applied in local or global coordination system"""
        if cosys is None:
            raise ValueError
        if cosys == 1:
            return "local"
        else:
            return "global"

    @staticmethod
    def _read_calcnode(cnode: int) -> str:
        """Reads if the calculation node was generated for the load record."""
        if cnode is None:
            raise ValueError
        if cnode == 1:
            return "generated"
        else:
            return "no"

    @staticmethod
    def _read_relabs(rela: int) -> str:
        """Reads relative or absolute coordination system parameter in Load record."""
        if rela is None:
            raise ValueError
        if rela == 1:
            return "relative"
        else:
            return "absolute"

    @staticmethod
    def _check_type(rec: rbt.IRobotLoadRecord) -> bool:
        # print(rec.Type)
        # print(type(rec.Type))
        # if type(rec.Type) is not int:
        #     raise ValueError
        rec_type = int(rec.Type)
        if rec_type in supported_load_types:
            return True
        else:
            return False

    @staticmethod
    def is_comb_nonlinear(lcase: rbt.IRobotCase) -> int:
        """Checks if combination/case is nonlinear type.
        Parameters: lcase: IRobotCase
        Returns: int"""
        if int(lcase.AnalizeType) == -1 or int(lcase.AnalizeType) == -3:
            return 1
        else:
            return 0

    def _read_contour_points(self, rec: rbt.IRobotLoadRecord) -> list[list, tuple]:
        """Reads contour points.
        Parameters: rec: IRobotLoadRecord
        Returns: list[contour_points, vector]"""
        npoints = int(rec.GetValue(13))
        rec_contour = rbt.IRobotLoadRecordInContour(rec)
        contour_points = []
        for p in range(1, npoints + 1):
            contour_points.append(rec_contour.GetContourPoint(p))
        vector = rec_contour.GetVector()
        return [contour_points, vector]

    def _is_3points(self, rec: rbt.IRobotLoadRecord) -> bool:
        """Checks if load on contour is variable type with 3 points
        Parameters: rec: IRobotLoadRecord
        Returns: boolean"""
        rec_contour = rbt.IRobotLoadRecordInContour(rec)
        points = [rec_contour.GetPoint(1), rec_contour.GetPoint(2), rec_contour.GetPoint(3)]
        # print(points)
        coord_list = []
        for coord in points:
            coord_list.append(coord[1])
            coord_list.append(coord[2])
            coord_list.append(coord[3])
        return any(coord_list)

    def _write_contour_points(self, points: list, column_index: int, contour_index: int) -> None:
        """Writes contour coordinates
        Parameters: points: list[tuples(x,y,z)]
                    loadnumber: int
                    column_index: int"""
        point_number = 1
        row = 3
        header = "Contour index: "
        self.ws_points[get_column_letter(column_index) + "1"] = header
        self.ws_points[get_column_letter(column_index + 1) + "1"] = contour_index
        self.ws_points[get_column_letter(column_index) + "2"] = "Point number"
        self.ws_points[get_column_letter(column_index + 1) + "2"] = "X"
        self.ws_points[get_column_letter(column_index + 2) + "2"] = "Y"
        self.ws_points[get_column_letter(column_index + 3) + "2"] = "Z"
        for point in points:
            self.ws_points[get_column_letter(column_index) + str(row)] = point_number
            self.ws_points[get_column_letter(column_index + 1) + str(row)] = point[0]
            self.ws_points[get_column_letter(column_index + 2) + str(row)] = point[1]
            self.ws_points[get_column_letter(column_index + 3) + str(row)] = point[2]
            point_number += 1
            row += 1

    def _write_3p_load_points(self, rec: rbt.IRobotLoadRecord, column_index: int) -> None:
        """Writes 3p coordinates A,B,C for P1, P2, P3 load distribution
        Parameters: rec: IRobotLoadRecord
                    column_index: int"""
        rec_contour = rbt.IRobotLoadRecordInContour(rec)
        pointA = "A" + str(rec_contour.GetPoint(1))
        pointB = "B" + str(rec_contour.GetPoint(2))
        pointC = "C" + str(rec_contour.GetPoint(3))
        self.ws_points[get_column_letter(column_index + 4) + "2"] = "3p definition"
        self.ws_points[get_column_letter(column_index + 4) + "3"] = pointA
        self.ws_points[get_column_letter(column_index + 4) + "4"] = pointB
        self.ws_points[get_column_letter(column_index + 4) + "5"] = pointC

    def _read_objects(self, rec: rbt.IRobotLoadRecord) -> list:
        """Reads the objects to which the load is applied to. For example bar numbers.
        Parameters: rec: IRobotLoadRecord"""
        objects = []
        rect = int(rec.Type)
        # print(rect)
        if rect != 89:  # if not a Body force
            if rect == 69:  # in edge load case objects are empty but do contain text
                obj = rec.Objects
                objects.append(obj.ToText())
                # else:
                #     objects.append(missing_msg)
            if rect == 7:  # in edge load case objects are empty but do contain text
                if rec.GetValue(15) == 1:
                    objects.append("all")
                else:
                    obj = rec.Objects
                    objects.append(obj.ToText())
                # else:
                #     objects.append(missing_ms
            else:
                count = rec.Objects.Count
                for r in range(1, count + 1):
                    objects.append(rec.Objects.Get(r))
        else:
            if rec.GetValue(12) == 1:
                objects.append(
                    "converted from body force masses! remove this row, convert masses to loads!"
                )  # body force is not supported, the load will be converted to self-weight during export
            else:
                # if count more than 0 then read text objects
                objects.append("converted from body force objects, insert objects manually")
        return objects

    def _write_load(
        self, lcase: rbt.IRobotCase, rec: rbt.IRobotLoadRecord, row: int, column_index: int, contour_index: int
    ) -> None:
        """Auxillary function, reads the load from the model, stores it in pyLoad object
        and returns them in list
        Parameters: lcase: IRobotCase,
                    rec: IRobotLoadRecord,
                    row: int,
                    column_index: int"""
        # load = pyLoad()
        rec_type = int(rec.Type)  # cast it as a int
        self.ws["B" + str(row)] = lcase.Name
        self.ws["A" + str(row)] = lcase.Number
        self.ws["H" + str(row)] = supported_load_types[rec_type]
        self.ws["I" + str(row)] = utilities.list_to_str(self._read_objects(rec))  # read the objects load is assigned to
        if rec_type == 0 or rec_type == 3:  # nodal force or point load on a bar
            self.ws["J" + str(row)] = round(rec.GetValue(0) / U, R)  # I_NFIPRV_FX
            self.ws["K" + str(row)] = round(rec.GetValue(1) / U, R)  # I_NFIPRV_FY
            self.ws["L" + str(row)] = round(rec.GetValue(2) / U, R)  # I_NFIPRV_FZ
            self.ws["M" + str(row)] = round(rec.GetValue(3) / U, R)  # I_NFIPRV_MX
            self.ws["N" + str(row)] = round(rec.GetValue(4) / U, R)  # I_NFIPRV_MY
            self.ws["O" + str(row)] = round(rec.GetValue(5) / U, R)  # I_NFIPRV_MZ
            self.ws["P" + str(row)] = rec.GetValue(8) * rad_to_deg  # I_NFIPRV_ALPHA
            self.ws["Q" + str(row)] = rec.GetValue(9) * rad_to_deg  # I_NFIPRV_BETA
            self.ws["R" + str(row)] = rec.GetValue(10) * rad_to_deg  # I_NFIPRV_GAMMA
            # load.calcnode = self.read_calcnode(rec.GetValue(12))
        if rec_type == 5:  # uniform load
            self.ws["J" + str(row)] = round(rec.GetValue(0) / U, R)
            self.ws["K" + str(row)] = round(rec.GetValue(1) / U, R)
            self.ws["L" + str(row)] = round(rec.GetValue(2) / U, R)
            self.ws["M" + str(row)] = round(rec.GetValue(3) / U, R)
            self.ws["N" + str(row)] = round(rec.GetValue(4) / U, R)
            self.ws["O" + str(row)] = round(rec.GetValue(5) / U, R)
            self.ws["P" + str(row)] = rec.GetValue(8) * rad_to_deg
            self.ws["Q" + str(row)] = rec.GetValue(9) * rad_to_deg
            self.ws["R" + str(row)] = rec.GetValue(10) * rad_to_deg
            self.ws["W" + str(row)] = self._read_cosystem(rec.GetValue(11))
            # load.projected = rec.GetValue(12)
            self.ws["X" + str(row)] = self._read_relabs(rec.GetValue(13))
            self.ws["T" + str(row)] = rec.GetValue(21)
            self.ws["U" + str(row)] = rec.GetValue(22)
        elif rec_type == 26:  # uniform load on a FE element
            self.ws["J" + str(row)] = round(rec.GetValue(0) / U, R)  # I_URV_PX
            self.ws["K" + str(row)] = round(rec.GetValue(1) / U, R)  # I_URV_PY
            self.ws["L" + str(row)] = round(rec.GetValue(2) / U, R)  # I_URV_PZ
            self.ws["W" + str(row)] = self._read_cosystem(rec.GetValue(11))  # I_URV_LOCAL_SYSTEM
            self.ws["Y" + str(row)] = rec.GetValue(12)  # I_URV_PROJECTION
        elif rec_type == 7:  # self-weight
            load_vector_X = rec.GetValue(0)
            load_vector_Y = rec.GetValue(1)
            load_vector_Z = rec.GetValue(2)
            factor = rec.GetValue(3)
            if load_vector_X != 0:
                self.ws["J" + str(row)] = factor * (math.copysign(1, load_vector_X))
                self.ws["K" + str(row)] = 0
                self.ws["L" + str(row)] = 0
            if load_vector_Y != 0:
                self.ws["J" + str(row)] = 0
                self.ws["K" + str(row)] = factor * (math.copysign(1, load_vector_Y))
                self.ws["L" + str(row)] = 0
            if load_vector_Z != 0:
                self.ws["J" + str(row)] = 0
                self.ws["K" + str(row)] = 0
                self.ws["L" + str(row)] = factor * (math.copysign(1, load_vector_Z))
        elif rec_type == 3:  # point load on a bar
            self.ws["W" + str(row)] = self._read_cosystem(rec.GetValue(11))  # I_BFCRV_LOC
            self.ws["V" + str(row)] = self._read_calcnode(rec.GetValue(12))  # I_BFCRV_GENERATE_CALC_NODE
            self.ws["X" + str(row)] = self._read_relabs(rec.GetValue(13))  # I_BFCRV_REL
            self.ws["S" + str(row)] = rec.GetValue(6)  # I_BFCRV_X
            self.ws["T" + str(row)] = rec.GetValue(21)  # I_BFCRV_OFFSET_Y
            self.ws["U" + str(row)] = rec.GetValue(22)  # I_BFCRV_OFFSET_Z
        elif rec_type == 6:  # "trapezoidal load (2p)"
            self.ws["M" + str(row)] = round(rec.GetValue(3) / U, R)  # I_BTRV_PX1
            self.ws["N" + str(row)] = round(rec.GetValue(4) / U, R)  # I_BTRV_PY1
            self.ws["O" + str(row)] = round(rec.GetValue(5) / U, R)  # I_BTRV_PZ1
            self.ws["J" + str(row)] = round(rec.GetValue(0) / U, R)  # I_BTRV_PX2
            self.ws["K" + str(row)] = round(rec.GetValue(1) / U, R)  # I_BTRV_PY2
            self.ws["L" + str(row)] = round(rec.GetValue(2) / U, R)  # I_BTRV_PZ2
            self.ws["T" + str(row)] = rec.GetValue(7)  # I_BTRV_X1
            self.ws["S" + str(row)] = rec.GetValue(6)  # I_BTRV_X2
            self.ws["P" + str(row)] = rec.GetValue(8) * rad_to_deg  # I_BTRV_ALPHA
            self.ws["Q" + str(row)] = rec.GetValue(9) * rad_to_deg  # I_BTRV_BETA
            self.ws["R" + str(row)] = rec.GetValue(10) * rad_to_deg  # I_BTRV_GAMMA
            self.ws["Y" + str(row)] = rec.GetValue(12)  # I_BTRV_PROJECTION
            self.ws["X" + str(row)] = self._read_relabs(rec.GetValue(13))  # I_BTRV_RELATIVE
        elif rec_type == 69:  # (FE) 2 load on edges
            self.ws["J" + str(row)] = round(rec.GetValue(0) / U, R)  # I_LOERV_PX
            self.ws["K" + str(row)] = round(rec.GetValue(1) / U, R)  # I_LOERV_PY
            self.ws["L" + str(row)] = round(rec.GetValue(2) / U, R)  # I_LOERV_PZ
            self.ws["M" + str(row)] = round(rec.GetValue(3) / U, R)  # I_LOERV_MX
            self.ws["N" + str(row)] = round(rec.GetValue(4) / U, R)  # I_LOERV_MY
            self.ws["O" + str(row)] = round(rec.GetValue(5) / U, R)  # I_LOERV_MZ
            self.ws["R" + str(row)] = rec.GetValue(6)  # I_LOERV_GAMMA
            self.ws["W" + str(row)] = self._read_cosystem(rec.GetValue(11))  # I_LOERV_LOCAL_SYSTEM
        elif rec_type == 89:  # "Body forces"
            # this load type is not fully supported, due to poor Robot API
            # it gets converted to self-weight, only one record for one direction
            self.messagebox(
                "Warning",
                "Body force load is not fully supported. It will be converted to self-weight, relative. Imported values need verification!",
            )
            relabs = rec.GetValue(13)
            # self.ws["X" + str(row)] = self._read_relabs(relabs)  # not needed for self-weight conversion
            if relabs == 1:
                conversion = 9.81
            else:
                conversion = 1
            load_vector_X = rec.GetValue(0)
            load_vector_Y = rec.GetValue(1)
            load_vector_Z = rec.GetValue(2)
            if load_vector_X != 0:
                self.ws["J" + str(row)] = load_vector_X / conversion
                self.ws["K" + str(row)] = 0
                self.ws["L" + str(row)] = 0
            elif load_vector_Y != 0:
                self.ws["J" + str(row)] = 0
                self.ws["K" + str(row)] = load_vector_Y / conversion
                self.ws["L" + str(row)] = 0
            elif load_vector_Z != 0:
                self.ws["J" + str(row)] = 0
                self.ws["K" + str(row)] = 0
                self.ws["L" + str(row)] = load_vector_Z / conversion
        elif rec_type == 28:  # "Load on contour"
            self.ws["M" + str(row)] = round(rec.GetValue(0) / U, R)  # I_ICRV_PX1
            self.ws["N" + str(row)] = round(rec.GetValue(1) / U, R)  # I_ICRV_PY1
            self.ws["O" + str(row)] = round(rec.GetValue(2) / U, R)  # I_ICRV_PZ1
            self.ws["J" + str(row)] = round(rec.GetValue(3) / U, R)  # I_ICRV_PX2
            self.ws["K" + str(row)] = round(rec.GetValue(4) / U, R)  # I_ICRV_PY2
            self.ws["L" + str(row)] = round(rec.GetValue(5) / U, R)  # I_ICRV_PZ2
            self.ws["S" + str(row)] = round(rec.GetValue(6) / U, R)  # I_ICRV_PX3
            self.ws["T" + str(row)] = round(rec.GetValue(7) / U, R)  # I_ICRV_PY3
            self.ws["U" + str(row)] = round(rec.GetValue(8) / U, R)  # I_ICRV_PZ3
            self.ws["Y" + str(row)] = rec.GetValue(12)  # I_ICRV_PROJECTION
            self.ws["W" + str(row)] = self._read_cosystem(rec.GetValue(15))  # I_ICRV_LOCAL
            points, vector = self._read_contour_points(rec=rec)
            self.ws["P" + str(row)] = vector[0]  # vector x coordinate
            self.ws["Q" + str(row)] = vector[1]  # vector y coordinate
            self.ws["R" + str(row)] = vector[2]  # vector z coordinate
            self.ws["Z" + str(row)] = contour_index  # Contour idenficator
            self._write_contour_points(points, row, column_index, contour_index)
            if self._is_3points(rec):
                self.ws["H" + str(row)] = "load 3p on contour"
                self._write_3p_load_points(rec, column_index)

    def _import_loadcases(self, cases: rbt.IRobotCaseServer) -> None:
        """Writes loadcases to the excel sheet.
        Parameters: cases: IRobotCaseServer"""
        ws_cases = self.wb[cases_sheet_name]
        row = 7
        for i in range(1, cases.Count + 1):  # loop1
            lcase = rbt.IRobotCase(cases.Get(i))
            if lcase.Type == rbt.IRobotCaseType.I_CT_SIMPLE:
                ws_cases["A" + str(row)] = lcase.Number
                ws_cases["B" + str(row)] = lcase.Name
                ws_cases["C" + str(row)] = supported_cases_nature[int(lcase.Nature)]
                # ws_cases["D" + str(row)] = int(lcase.Nature)
                case = rbt.IRobotSimpleCase(lcase)
                if case.IsAuxiliary:
                    ws_cases["H" + str(row)] = 1
                else:
                    ws_cases["H" + str(row)] = 0
                if case.AnalizeType == rbt.IRobotCaseAnalizeType.I_CAT_STATIC_LINEAR:
                    ws_cases["D" + str(row)] = 0
                    ws_cases["E" + str(row)] = 1
                    ws_cases["F" + str(row)] = 0
                    ws_cases["G" + str(row)] = 0
                elif lcase.AnalizeType == rbt.IRobotCaseAnalizeType.I_CAT_STATIC_NONLINEAR:
                    ws_cases["D" + str(row)] = 1
                    ws_cases["E" + str(row)] = 2
                    params = rbt.IRobotNonlinearAnalysisParams(case.GetAnalysisParams())
                    if params.MatrixUpdateAfterEachIteration:
                        ws_cases["F" + str(row)] = 1
                    else:
                        ws_cases["F" + str(row)] = 0
                    if params.PDelta:
                        ws_cases["G" + str(row)] = 1
                    else:
                        ws_cases["G" + str(row)] = 0
                elif lcase.AnalizeType == rbt.IRobotCaseAnalizeType.I_CAT_STATIC_BUCKLING:
                    ws_cases["D" + str(row)] = 0
                    ws_cases["E" + str(row)] = 4
                    params = rbt.IRobotBucklingAnalysisParams(case.GetAnalysisParams())
                    if params.MatrixUpdateAfterEachIteration:
                        ws_cases["F" + str(row)] = 1
                    else:
                        ws_cases["F" + str(row)] = 0
                    if params.PDelta:
                        ws_cases["G" + str(row)] = 1
                    else:
                        ws_cases["G" + str(row)] = 0
                row += 1

    def _import_cases_in_combination_sheet(self, cases: rbt.IRobotCaseServer) -> None:
        """Imports cases in top rows in combinations sheet"""
        ws_comb = self.wb[combinations_sheet_name]
        col_index = 8
        for i in range(1, cases.Count + 1):  # loop1
            lcase = rbt.IRobotCase(cases.Get(i))
            # Propagating loadcases in combinations sheet
            col_letter = get_column_letter(col_index)
            # address_name = col_letter + "3"
            ws_comb[col_letter + "3"].value = lcase.Name
            ws_comb[col_letter + "4"].value = lcase.Number
            col_index += 1

    def _propagate_factors(self, ws: Worksheet, lcomb: rbt.IRobotCaseCombination, row_index: int) -> None:
        """Auxillary function to propagate factors for load combinations
        Parameters: ws: Workbook
                    lcomb: IRobotCaseCombination
                    row_index: int"""
        # max column index containg cases
        max_column_index = utilities.max_column_index(ws, min_column=8, min_row=4, max_row=4)
        # get column letter
        col_letter = get_column_letter(max_column_index)
        columns_range = "H4:" + col_letter + "4"
        # getting lcombination components
        case_factor_mng = lcomb.CaseFactors
        for n in range(1, case_factor_mng.Count + 1):
            case_factor = case_factor_mng.Get(n)
            # comb_number = lcomb.Number # never used
            case_number = case_factor.CaseNumber
            case_factor = case_factor.Factor
            for row in ws[columns_range]:
                for cell in row:
                    if int(cell.value) == case_number:
                        matching_column = str(cell.column_letter)
            address = matching_column + str(row_index)
            ws[address].value = case_factor

    def _import_combinations(self, cases: rbt.IRobotCaseServer) -> None:
        """Writes combinations to the excel sheet.
        Parameters: cases: IRobotCaseServer"""
        ws_comb = self.wb[combinations_sheet_name]
        row = 7
        for i in range(1, cases.Count + 1):  # loop2
            lcase = rbt.IRobotCase(cases.Get(i))
            if int(lcase.Type) == 1:
                lcomb = rbt.IRobotCaseCombination(lcase)
                ws_comb["A" + str(row)] = lcomb.Number
                ws_comb["B" + str(row)] = lcomb.Name
                ws_comb["C" + str(row)] = combination_type[int(lcomb.CombinationType)]
                ws_comb["D" + str(row)] = self.is_comb_nonlinear(lcomb)
                if self.is_comb_nonlinear(lcomb):
                    params = rbt.IRobotNonlinearAnalysisParams(lcomb.GetAnalysisParams())
                    if params.MatrixUpdateAfterEachIteration:
                        ws_comb["F" + str(row)] = 1
                    if params.PDelta:
                        ws_comb["G" + str(row)] = 1
                ws_comb["E" + str(row)] = int(lcomb.AnalizeType)
                self._propagate_factors(ws=ws_comb, lcomb=lcomb, row_index=row)
                row += 1

    def _import_load(self, cases: rbt.IRobotCaseServer) -> None:
        """Imports loads from Robot model to the excel spreadsheet
        Parameters: cases: IRobotCaseServer"""
        start_row = 8
        column_index = 1  # index of the column used for contour loads'
        contour_index = 1  # contour number used for identification
        for i in range(1, cases.Count + 1):  # loop3
            lcase = rbt.IRobotCase(cases.Get(i))
            if lcase.AnalizeType in supported_analize_type:
                if lcase.Type == rbt.IRobotCaseType.I_CT_SIMPLE:
                    lsimplecase = rbt.IRobotSimpleCase(lcase)
                    for r in range(1, lsimplecase.Records.Count + 1):
                        # Get load record
                        rec = lsimplecase.Records.Get(r)
                        # Check if load record is not auto generated from claddings loads
                        rec_com = rbt.IRobotLoadRecordCommon(lsimplecase.Records.Get(r))
                        if not rec_com.IsAutoGenerated:
                            # check if record is supported
                            if self._check_type(rec):
                                self._write_load(
                                    lcase=lsimplecase,
                                    rec=rec,
                                    row=start_row,
                                    column_index=column_index,
                                    contour_index=contour_index,
                                )
                                if int(rec.Type) == 28:
                                    column_index += 6
                                    contour_index += 1
                                start_row += 1

    def import_loads_and_comb(self, import_loads: bool, import_comb: bool) -> None:
        "Main function, imports loads and combinations to Excel sheet."
        start_time = time.time()
        self.ws = self.wb[load_sheet_name]
        # Clear the range for loadcases, which are always imported
        utilities.clear_range2(
            self.wb[cases_sheet_name], min_row=7, max_row=1000, min_col=1, max_col=9
        )  # Clearing cases
        # Get all loadcases from the model
        cases = self.structure.Cases.GetAll()
        print("Getting all cases " + str(time.time() - start_time))
        # Import loadcases, this is always executed
        self._import_loadcases(cases)
        print("Import loadcses " + str(time.time() - start_time))
        # Import combinations
        if import_comb == 1:
            utilities.clear_range2(
                self.wb[combinations_sheet_name], min_row=3, max_row=4, min_col=8, max_col=500
            )  # Clearing combinations upper row
            utilities.clear_range2(
                self.wb[combinations_sheet_name], min_row=7, max_row=476, min_col=1, max_col=500
            )  # Clearing combinations factors
            self._import_cases_in_combination_sheet(cases)
            self._import_combinations(cases)
            print("Import combinations " + str(time.time() - start_time))
        if import_loads == 1:
            utilities.clear_range2(self.ws, min_row=8, max_row=1000, min_col=1, max_col=2)  # Clearing loads
            utilities.clear_range2(self.ws, min_row=8, max_row=1000, min_col=8, max_col=25)  # Clearing loads
            utilities.clear_range2(
                self.wb[points_sheet_name], min_row=1, max_row=500, min_col=1, max_col=500
            )  # Clearing load contour points
            self._import_load(cases)
            print("Import loads " + str(time.time() - start_time))
            print("Finished.")
        self.wb.save(self.path)
